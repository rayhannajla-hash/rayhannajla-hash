"""
Task 2: Sortir email di Excel berdasarkan penerima langsung.
Baca sheet harian, kelompokkan per kategori, tulis ke sheet 'Sorted'.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
import config

CATEGORY_ORDER = ["Direct To", "To (Banyak)", "CC", "Broadcast", "Lainnya"]

# (warna teks badge, warna bg section header, warna bg baris data)
CATEGORY_STYLE = {
    "Direct To":   ("FFFFFF", "1E8449", "EAFAF1"),  # hijau
    "To (Banyak)": ("FFFFFF", "1A5276", "EBF5FB"),  # biru
    "CC":          ("333333", "D4AC0D", "FEFDE7"),  # kuning
    "Broadcast":   ("FFFFFF", "7D3C98", "F5EEF8"),  # ungu
    "Lainnya":     ("FFFFFF", "616A6B", "F2F3F4"),  # abu
}

THIN = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)


def read_today_emails(wb):
    sheet_name = datetime.now().strftime("%Y-%m-%d")
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' tidak ditemukan. "
            "Jalankan Task 1 (email_fetcher.py) terlebih dahulu."
        )
    ws      = wb[sheet_name]
    headers = [c.value for c in ws[2]]   # baris 2 = header kolom
    emails  = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        emails.append(dict(zip(headers, row)))
    return emails


def group_by_category(emails):
    groups = {cat: [] for cat in CATEGORY_ORDER}
    for email in emails:
        cat = email.get("Kategori", "Lainnya")
        groups.setdefault(cat, []).append(email)
    return groups


def _col_widths(ws):
    widths = {"A": 5, "B": 10, "C": 10, "D": 44, "E": 24,
              "F": 28, "G": 22, "H": 12, "I": 55}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _section_banner(ws, row, label, count, header_bg):
    """Baris judul seksi berwarna penuh."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    c = ws.cell(row=row, column=1)
    icon = {"Direct To": "🟢", "To (Banyak)": "🔵", "CC": "🟡",
            "Broadcast": "🟣", "Lainnya": "⚪"}.get(label, "▪")
    c.value     = f"  {icon}  {label.upper()}  —  {count} email"
    c.font      = Font(bold=True, color="FFFFFF", size=11)
    c.fill      = PatternFill("solid", fgColor=header_bg)
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 24
    return row + 1


def _col_header(ws, row, header_bg):
    headers = ["No", "Jam", "Tanggal", "Subject", "Dari",
               "Kepada (To)", "Alamat Pengirim", "Kategori", "Preview Isi"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font      = Font(bold=True, color="FFFFFF", size=9)
        c.fill      = PatternFill("solid", fgColor="2C3E50")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = THIN
    ws.row_dimensions[row].height = 18
    return row + 1


def _data_rows(ws, emails, start_row, row_bg):
    fill = PatternFill("solid", fgColor=row_bg)
    for i, email in enumerate(emails, 1):
        row    = start_row + i - 1
        values = [
            i,
            email.get("Jam", ""),
            email.get("Tanggal", ""),
            email.get("Subject", ""),
            email.get("Dari", ""),
            email.get("Kepada (To)", ""),
            email.get("Alamat Pengirim", ""),
            email.get("Kategori", ""),
            email.get("Preview Isi", ""),
        ]
        for col, val in enumerate(values, 1):
            c           = ws.cell(row=row, column=col, value=val)
            c.fill      = fill
            c.border    = THIN
            c.alignment = Alignment(vertical="center", wrap_text=(col == 9))
        ws.row_dimensions[row].height = 17
    return start_row + len(emails)


def build_sorted_sheet(wb, groups):
    sheet_name = "Sorted_" + datetime.now().strftime("%Y-%m-%d")
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    _col_widths(ws)

    # Judul utama
    today_str = datetime.now().strftime("%A, %d %B %Y")
    total     = sum(len(v) for v in groups.values())
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value     = f"📊  EMAIL TERSORTING  |  {today_str}  |  Target: {config.TARGET_NAME}  |  Total: {total}"
    c.font      = Font(bold=True, color="FFFFFF", size=13)
    c.fill      = PatternFill("solid", fgColor="0A2342")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    current_row = 3
    stats       = {}
    for cat in CATEGORY_ORDER:
        emails = groups.get(cat, [])
        stats[cat] = len(emails)
        if not emails:
            continue
        _, header_bg, row_bg = CATEGORY_STYLE[cat]
        current_row = _section_banner(ws, current_row, cat, len(emails), header_bg)
        current_row = _col_header(ws, current_row, header_bg)
        current_row = _data_rows(ws, emails, current_row, row_bg)
        current_row += 1  # baris kosong antar seksi

    ws.freeze_panes = "A2"
    return ws, sheet_name, stats


def run():
    print("=" * 55)
    print("TASK 2 — Sortir Email by Recipient")
    print("=" * 55)

    filepath = Path(config.EXCEL_PATH)
    if not filepath.exists():
        print("  File tidak ditemukan. Jalankan Task 1 terlebih dahulu.")
        return

    wb     = openpyxl.load_workbook(filepath)
    emails = read_today_emails(wb)
    print(f"  Membaca {len(emails)} email...")

    groups = group_by_category(emails)
    for cat in CATEGORY_ORDER:
        count = len(groups.get(cat, []))
        if count:
            print(f"    {cat:20s}: {count} email")

    _, sheet_name, _ = build_sorted_sheet(wb, groups)
    wb.save(filepath)
    print(f"  Sheet '{sheet_name}' selesai.")
    print(f"  File: {filepath}")


if __name__ == "__main__":
    run()
