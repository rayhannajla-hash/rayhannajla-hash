"""
Task 1: Copy email baru dari Outlook ke Excel
Baca inbox Outlook via win32com, simpan ke sheet harian di Excel.
"""

import win32com.client
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from pathlib import Path
import re
import config


def get_outlook_inbox():
    outlook = win32com.client.Dispatch("Outlook.Application")
    namespace = outlook.GetNamespace("MAPI")
    return namespace.GetDefaultFolder(6)  # 6 = olFolderInbox


def parse_recipients(recipients_obj):
    addresses = []
    for recipient in recipients_obj:
        addresses.append({
            "address": recipient.Address.lower().strip(),
            "name": recipient.Name.strip(),
            "type": recipient.Type,  # 1=To, 2=CC, 3=BCC
        })
    return addresses


def classify_email(recipients):
    target_name  = config.TARGET_NAME.lower()
    target_email = config.TARGET_EMAIL.lower()

    to_list = [r for r in recipients if r["type"] == 1]
    cc_list = [r for r in recipients if r["type"] == 2]

    for r in to_list:
        if target_email in r["address"] or target_name in r["name"].lower():
            return "Direct To" if len(to_list) == 1 else "To (Banyak)"

    for r in cc_list:
        if target_email in r["address"] or target_name in r["name"].lower():
            return "CC"

    if len(to_list) > config.BROADCAST_THRESHOLD:
        return "Broadcast"

    return "Lainnya"


def fetch_emails_since(hours_back=None):
    hours_back = hours_back or config.FETCH_HOURS_BACK
    since_dt   = datetime.now() - timedelta(hours=hours_back)

    inbox    = get_outlook_inbox()
    messages = inbox.Items
    messages.Sort("[ReceivedTime]", True)

    emails = []
    for msg in messages:
        try:
            received      = msg.ReceivedTime.replace(tzinfo=None)
            if received < since_dt:
                break
            recipients    = parse_recipients(msg.Recipients)
            to_names      = "; ".join(r["name"] for r in recipients if r["type"] == 1)
            cc_names      = "; ".join(r["name"] for r in recipients if r["type"] == 2)
            body_preview  = re.sub(r"\s+", " ", msg.Body or "").strip()
            body_preview  = body_preview[:config.BODY_PREVIEW_LENGTH]
            emails.append({
                "received":     received,
                "subject":      (msg.Subject or "").strip(),
                "sender":       msg.SenderName or "",
                "sender_email": msg.SenderEmailAddress or "",
                "to":           to_names,
                "cc":           cc_names,
                "category":     classify_email(recipients),
                "body_preview": body_preview,
            })
        except Exception as e:
            print(f"  [skip] {e}")
    return emails


# ── Styling helpers ────────────────────────────────────────────────────────────

HEADER_BG   = "1B4F72"   # biru tua — baris header kolom
TITLE_BG    = "0A2342"   # biru sangat tua — judul sheet
SUBROW_BG   = "EBF5FB"   # biru sangat muda — baris alternating

CAT_COLOR = {
    "Direct To":   "2ECC71",   # hijau   — paling penting
    "To (Banyak)": "85C1E9",   # biru muda
    "CC":          "F7DC6F",   # kuning
    "Broadcast":   "D7BDE2",   # ungu muda
    "Lainnya":     "BFC9CA",   # abu
}

CAT_ROW_BG = {
    "Direct To":   "EAFAF1",
    "To (Banyak)": "EBF5FB",
    "CC":          "FEFDE7",
    "Broadcast":   "F5EEF8",
    "Lainnya":     "F2F3F4",
}

THIN = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)


def _col_widths(ws):
    widths = {"A": 5, "B": 10, "C": 10, "D": 44, "E": 24,
              "F": 28, "G": 28, "H": 22, "I": 12, "J": 55}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _write_title(ws, today_str, total):
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value      = f"📥  LOG EMAIL HARIAN  |  {today_str}  |  Total: {total} email"
    c.font       = Font(bold=True, color="FFFFFF", size=13)
    c.fill       = PatternFill("solid", fgColor=TITLE_BG)
    c.alignment  = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30


def _write_col_header(ws, row=2):
    headers = ["No", "Jam", "Tanggal", "Subject", "Dari",
               "Kepada (To)", "CC", "Alamat Pengirim", "Kategori", "Preview Isi"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", fgColor=HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = THIN
    ws.row_dimensions[row].height = 22


def _write_data_rows(ws, emails, start_row=3):
    for i, email in enumerate(emails, 1):
        row    = start_row + i - 1
        cat    = email["category"]
        bg     = CAT_ROW_BG.get(cat, "FFFFFF")
        fill   = PatternFill("solid", fgColor=bg)
        values = [
            i,
            email["received"].strftime("%H:%M"),
            email["received"].strftime("%d/%m/%Y"),
            email["subject"],
            email["sender"],
            email["to"],
            email["cc"],
            email["sender_email"],
            cat,
            email["body_preview"],
        ]
        for col, val in enumerate(values, 1):
            c           = ws.cell(row=row, column=col, value=val)
            c.fill      = fill
            c.border    = THIN
            c.alignment = Alignment(vertical="center", wrap_text=(col == 10))
            if col == 9:  # badge kategori
                badge_color = CAT_COLOR.get(cat, "BFC9CA")
                c.font = Font(bold=True, color=badge_color, size=10)
        ws.row_dimensions[row].height = 17


def get_or_create_workbook(filepath):
    if Path(filepath).exists():
        return openpyxl.load_workbook(filepath)
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    return wb


def run():
    print("=" * 55)
    print("TASK 1 — Copy Email dari Outlook ke Excel")
    print("=" * 55)

    filepath = Path(config.EXCEL_PATH)
    filepath.parent.mkdir(parents=True, exist_ok=True)

    print(f"  Mengambil email {config.FETCH_HOURS_BACK} jam terakhir...")
    emails = fetch_emails_since()
    print(f"  Ditemukan {len(emails)} email.")

    if not emails:
        print("  Tidak ada email baru. Selesai.")
        return 0

    wb         = get_or_create_workbook(filepath)
    today_str  = datetime.now().strftime("%A, %d %B %Y")
    sheet_name = datetime.now().strftime("%Y-%m-%d")

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    _col_widths(ws)
    _write_title(ws, today_str, len(emails))
    _write_col_header(ws)
    _write_data_rows(ws, emails)

    # Freeze baris header
    ws.freeze_panes = "A3"

    wb.save(filepath)
    print(f"  {len(emails)} email ditulis ke sheet '{sheet_name}'")
    print(f"  File: {filepath}")
    return len(emails)


if __name__ == "__main__":
    run()
