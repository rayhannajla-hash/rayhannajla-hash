"""
Task 3: Buat summary email setiap pagi.
Generate sheet Summary dengan dashboard statistik, topik, dan action items.
Opsional: kirim ringkasan ke Telegram via bot AINUN.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import datetime
from pathlib import Path
from collections import Counter
import re
import config

try:
    import requests
    REQUESTS_AVAILABLE = True
except ImportError:
    REQUESTS_AVAILABLE = False

CATEGORY_ORDER = ["Direct To", "To (Banyak)", "CC", "Broadcast", "Lainnya"]

CAT_COLOR = {
    "Direct To":   "1E8449",
    "To (Banyak)": "1A5276",
    "CC":          "D4AC0D",
    "Broadcast":   "7D3C98",
    "Lainnya":     "616A6B",
}

ACTION_KEYWORDS = [
    "mohon", "tolong", "please", "request", "urgent", "segera",
    "deadline", "due", "approval", "approve", "follow up", "reminder",
    "konfirmasi", "confirm", "review", "check", "update",
]

THIN = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)


# ── Baca data ─────────────────────────────────────────────────────────────────

def read_emails(wb):
    """Baca dari sheet Sorted, fallback ke sheet raw."""
    sorted_name = "Sorted_" + datetime.now().strftime("%Y-%m-%d")
    raw_name    = datetime.now().strftime("%Y-%m-%d")
    source = sorted_name if sorted_name in wb.sheetnames else raw_name
    if source not in wb.sheetnames:
        raise ValueError("Tidak ada data hari ini. Jalankan Task 1 & 2 dulu.")

    ws      = wb[source]
    # Cari baris header (ada teks "Subject")
    headers = None
    emails  = []
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            if row and "Subject" in (row[3] or ""):
                headers = list(row)
            continue
        if row[0] and headers:
            try:
                int(row[0])
                emails.append(dict(zip(headers, row)))
            except (TypeError, ValueError):
                continue
    return emails


# ── Analisis ──────────────────────────────────────────────────────────────────

def count_by_category(emails):
    counts = {cat: 0 for cat in CATEGORY_ORDER}
    for e in emails:
        cat = e.get("Kategori", "Lainnya")
        if cat in counts:
            counts[cat] += 1
    return counts


def top_senders(emails, n=5):
    return Counter(e.get("Dari", "") for e in emails).most_common(n)


def top_keywords(emails, n=8):
    stopwords = {
        "re", "fwd", "fw", "dari", "untuk", "dan", "atau", "yang", "dengan",
        "di", "ke", "the", "and", "or", "for", "is", "in", "on", "to",
        "a", "of", "by", "an", "be", "at", "as", "it", "we", "all",
    }
    counter = Counter()
    for e in emails:
        words = re.findall(r"[a-zA-Z]{3,}", (e.get("Subject") or "").lower())
        counter.update(w for w in words if w not in stopwords)
    return counter.most_common(n)


def find_action_items(emails):
    items = []
    for e in emails:
        text = ((e.get("Subject") or "") + " " + (e.get("Preview Isi") or "")).lower()
        matched = [kw for kw in ACTION_KEYWORDS if kw in text]
        if matched:
            items.append({
                "subject":  e.get("Subject", ""),
                "dari":     e.get("Dari", ""),
                "kategori": e.get("Kategori", ""),
                "keyword":  ", ".join(matched[:3]),
            })
    return items


# ── Excel helpers ─────────────────────────────────────────────────────────────

def _cell(ws, row, col, value="", bold=False, size=10, color="000000",
          bg=None, align="left", border=False, wrap=False, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, size=size, color=color, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center",
                            wrap_text=wrap, indent=(1 if align == "left" else 0))
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if border:
        c.border = THIN
    return c


def _section_title(ws, row, col, text, bg, colspan=5):
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + colspan - 1)
    _cell(ws, row, col, text, bold=True, size=11, color="FFFFFF",
          bg=bg, align="left")
    ws.row_dimensions[row].height = 22
    return row + 1


def _kv(ws, row, col, key, value, alt=False, key_width=None):
    bg = "F0F3F4" if alt else "FFFFFF"
    _cell(ws, row, col,     key,   bold=True,  size=10, bg=bg, border=True)
    _cell(ws, row, col + 1, value, bold=False, size=10, bg=bg, border=True)
    ws.row_dimensions[row].height = 17


def _bar_chart(ws, data_ws_title, cat_row, count_row, anchor):
    """Buat bar chart statistik kategori."""
    chart       = BarChart()
    chart.type  = "col"
    chart.title = "Email per Kategori"
    chart.style = 10
    chart.height = 10
    chart.width  = 16

    data   = Reference(ws, min_col=count_row[0], min_row=count_row[1],
                       max_col=count_row[2], max_row=count_row[3])
    cats   = Reference(ws, min_col=cat_row[0],   min_row=cat_row[1],
                       max_col=cat_row[2],        max_row=cat_row[3])
    chart.add_data(data)
    chart.set_categories(cats)
    chart.series[0].title = None
    ws.add_chart(chart, anchor)


# ── Build summary sheet ───────────────────────────────────────────────────────

def build_summary_sheet(wb, emails):
    sheet_name = "SUMMARY_" + datetime.now().strftime("%Y-%m-%d")
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, 0)  # taruh paling depan

    # Lebar kolom
    col_w = {"A": 3, "B": 24, "C": 18, "D": 3, "E": 24, "F": 18,
             "G": 3, "H": 24, "I": 18, "J": 3}
    for col, w in col_w.items():
        ws.column_dimensions[col].width = w

    now       = datetime.now()
    today_str = now.strftime("%A, %d %B %Y")
    counts    = count_by_category(emails)
    actions   = find_action_items(emails)
    direct    = [e for e in emails if e.get("Kategori") == "Direct To"]

    # ── HEADER UTAMA ──────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    _cell(ws, 1, 1,
          f"📊  DAILY EMAIL SUMMARY  |  {today_str}  |  {config.TARGET_NAME}",
          bold=True, size=14, color="FFFFFF", bg="0A2342", align="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:I2")
    _cell(ws, 2, 1,
          f"Generated: {now.strftime('%H:%M:%S')}   |   "
          f"Total Email Masuk: {len(emails)}   |   "
          f"Direct to {config.TARGET_NAME}: {counts.get('Direct To', 0)}   |   "
          f"Action Items: {len(actions)}",
          italic=True, size=10, color="5D6D7E", align="center")
    ws.row_dimensions[2].height = 18

    # ── KOLOM KIRI: Statistik & Top Sender ────────────────────────
    row = 4

    # Statistik per kategori
    row = _section_title(ws, row, 2, "  STATISTIK PER KATEGORI", "1B4F72", 2)
    for i, cat in enumerate(CATEGORY_ORDER):
        count = counts.get(cat, 0)
        pct   = f"{count/len(emails)*100:.0f}%" if emails else "0%"
        bg    = "F0F3F4" if i % 2 else "FFFFFF"
        _cell(ws, row, 2, f"  {cat}", bold=(cat == "Direct To"),
              size=10, color=CAT_COLOR.get(cat, "000000"), bg=bg, border=True)
        _cell(ws, row, 3, f"{count} email  ({pct})", size=10, bg=bg, border=True,
              bold=(cat == "Direct To"))
        ws.row_dimensions[row].height = 17
        row += 1
    # Total
    _cell(ws, row, 2, "  TOTAL", bold=True, size=10,
          color="FFFFFF", bg="1B4F72", border=True)
    _cell(ws, row, 3, f"{len(emails)} email", bold=True, size=10,
          color="FFFFFF", bg="1B4F72", border=True)
    ws.row_dimensions[row].height = 18
    row += 2

    # Top Senders
    row = _section_title(ws, row, 2, "  TOP PENGIRIM", "6C3483", 2)
    for i, (sender, count) in enumerate(top_senders(emails)):
        _kv(ws, row, 2, f"  {sender}", f"{count} email", alt=(i % 2 == 0))
        row += 1
    row += 1

    # ── KOLOM TENGAH: Topik & Action Items ────────────────────────
    row2 = 4
    row2 = _section_title(ws, row2, 5, "  KATA KUNCI TOPIK", "0E6655", 2)
    for i, (word, count) in enumerate(top_keywords(emails)):
        _kv(ws, row2, 5, f"  {word.title()}", f"{count}x", alt=(i % 2 == 0))
        row2 += 1
    row2 += 2

    row2 = _section_title(ws, row2, 5,
                          f"  ACTION ITEMS  ({len(actions)} email)",
                          "922B21", 2)
    if actions:
        for i, item in enumerate(actions[:12]):
            bg = "FEF9E7" if i % 2 == 0 else "FDEDEC"
            _cell(ws, row2, 5,
                  f"  {(item['subject'] or '')[:38]}",
                  size=9, bg=bg, border=True)
            _cell(ws, row2, 6,
                  f"{item['dari'][:18]} | {item['keyword']}",
                  size=9, bg=bg, border=True)
            ws.row_dimensions[row2].height = 16
            row2 += 1
    else:
        _cell(ws, row2, 5, "  Tidak ada action items terdeteksi",
              italic=True, size=10, color="7F8C8D")
        row2 += 1

    # ── KOLOM KANAN: Direct Email Detail ──────────────────────────
    row3 = 4
    row3 = _section_title(ws, row3, 8,
                          f"  DIRECT TO {config.TARGET_NAME.upper()} ({len(direct)})",
                          "1E8449", 2)
    if direct:
        for i, e in enumerate(direct[:20]):
            bg = "EAFAF1" if i % 2 == 0 else "FDFEFE"
            _cell(ws, row3, 8, f"  {e.get('Dari', '')[:22]}",
                  bold=True, size=9, bg=bg, border=True)
            _cell(ws, row3, 9, f"{(e.get('Subject') or '')[:28]}",
                  size=9, bg=bg, border=True)
            ws.row_dimensions[row3].height = 16
            row3 += 1
    else:
        _cell(ws, row3, 8, "  Tidak ada email direct hari ini",
              italic=True, size=10, color="7F8C8D")

    ws.freeze_panes = "B3"
    return ws, sheet_name


# ── Telegram ──────────────────────────────────────────────────────────────────

def telegram_message(emails, counts, actions):
    now    = datetime.now()
    direct = counts.get("Direct To", 0)
    lines  = [
        f"☀️ *DAILY EMAIL SUMMARY*",
        f"📅 {now.strftime('%A, %d %B %Y')}  |  {now.strftime('%H:%M')}",
        "",
        f"📧 Total Email : *{len(emails)}*",
        f"🎯 Direct to {config.TARGET_NAME} : *{direct}*",
        f"⚠️ Action Items : *{len(actions)}*",
        "",
        "*Per Kategori:*",
    ]
    icons = {"Direct To": "🟢", "To (Banyak)": "🔵",
             "CC": "🟡", "Broadcast": "🟣", "Lainnya": "⚪"}
    for cat in CATEGORY_ORDER:
        count = counts.get(cat, 0)
        if count:
            lines.append(f"{icons.get(cat, '▪')} {cat}: {count}")
    if actions:
        lines += ["", "*⚠️ Perlu Perhatian:*"]
        for item in actions[:5]:
            lines.append(f"• {(item['subject'] or '')[:50]}")
    return "\n".join(lines)


def send_to_telegram(message):
    if not REQUESTS_AVAILABLE:
        print("  [Telegram] install requests dulu: pip install requests")
        return
    if not config.TELEGRAM_BOT_TOKEN or not config.TELEGRAM_CHAT_ID:
        print("  [Telegram] Token/Chat ID belum diisi di config.py")
        return
    try:
        r = requests.post(
            f"https://api.telegram.org/bot{config.TELEGRAM_BOT_TOKEN}/sendMessage",
            json={"chat_id": config.TELEGRAM_CHAT_ID,
                  "text": message, "parse_mode": "Markdown"},
            timeout=10,
        )
        if r.status_code == 200:
            print("  [Telegram] ✅ Pesan terkirim.")
        else:
            print(f"  [Telegram] ❌ {r.status_code}: {r.text[:80]}")
    except Exception as e:
        print(f"  [Telegram] Error: {e}")


# ── Entry point ───────────────────────────────────────────────────────────────

def run():
    print("=" * 55)
    print("TASK 3 — Generate Summary Email Harian")
    print("=" * 55)

    filepath = Path(config.EXCEL_PATH)
    if not filepath.exists():
        print("  File tidak ditemukan. Jalankan Task 1 & 2 dulu.")
        return

    wb     = openpyxl.load_workbook(filepath)
    emails = read_emails(wb)
    print(f"  Memproses {len(emails)} email...")

    ws, sheet_name = build_summary_sheet(wb, emails)
    wb.save(filepath)
    print(f"  Sheet '{sheet_name}' selesai.")
    print(f"  File: {filepath}")

    if config.SEND_TELEGRAM:
        counts  = count_by_category(emails)
        actions = find_action_items(emails)
        msg     = telegram_message(emails, counts, actions)
        print("\n--- PREVIEW TELEGRAM ---")
        print(msg)
        print("------------------------")
        send_to_telegram(msg)


if __name__ == "__main__":
    run()
